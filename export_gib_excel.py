"""
GİB Excel Export Modülü
İndirilecek KDV Listesi için iki farklı Excel çıktısı üretir:
1. Özet Liste (fatura bazlı)
2. Kalem Bazlı Liste (satır bazlı)
"""

import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _get_styles():
    """Ortak Excel stilleri"""
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(bold=True, size=14, color="2E74B5")
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'title_font': title_font,
        'title_alignment': title_alignment,
        'thin_border': thin_border,
        'number_alignment': number_alignment
    }


def generate_gib_ozet_excel(invoices, output_stream=None):
    """
    GİB Özet Liste Excel dosyası oluştur.
    Her fatura = 1 satır.
    
    Args:
        invoices: Fatura listesi (dict listesi)
        output_stream: Opsiyonel BytesIO stream. None ise yeni oluşturulur.
        
    Returns:
        BytesIO: Excel dosyası içeren stream
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli: pip install openpyxl")
    
    if output_stream is None:
        output_stream = io.BytesIO()
    
    # Silinen faturaları filtrele
    active_invoices = [inv for inv in invoices if not inv.get('_deleted', False)]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GİB Özet Liste"
    
    styles = _get_styles()
    
    # Başlık satırı
    ws.merge_cells('A1:I1')
    ws['A1'] = "GİB İNDİRİLECEK KDV LİSTESİ - ÖZET"
    ws['A1'].font = styles['title_font']
    ws['A1'].alignment = styles['title_alignment']
    
    # Tarih bilgisi
    ws['A2'] = f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=9, color="666666")
    
    # Kolon başlıkları (Satır 4)
    headers = [
        "Fatura Tarihi",
        "Fatura No",
        "Satıcı VKN/TCKN",
        "Satıcı Unvan",
        "Mal/Hizmet Tutarı (KDV Hariç)",
        "KDV Oranı (%)",
        "Hesaplanan KDV",
        "İndirilecek KDV",
        "Belge Türü"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']
    
    # Kolon genişlikleri
    col_widths = [15, 22, 15, 40, 22, 12, 18, 18, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Veri satırları
    total_kdv_haric = 0
    total_kdv = 0
    total_indirilen = 0
    
    for row_idx, inv in enumerate(active_invoices, 5):
        # KDV oranı hesapla
        kdv_haric = float(inv.get('kdv_haric_tutar', 0) or 0)
        kdv = float(inv.get('kdv', 0) or 0)
        kdv_orani = round((kdv / kdv_haric * 100), 0) if kdv_haric > 0 else 0
        
        fatura_no = f"{inv.get('seri', '')}{inv.get('sira_no', '')}"
        
        row_data = [
            inv.get('tarih', ''),
            fatura_no,
            inv.get('satici_vkn', ''),
            inv.get('satici_unvan', ''),
            kdv_haric,
            kdv_orani,
            kdv,
            float(inv.get('toplam_indirilen_kdv', 0) or 0),
            "E-FATURA"
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['thin_border']
            
            # Sayısal kolonlar için format
            if col_idx in [5, 7, 8]:  # Tutar kolonları
                cell.number_format = '#,##0.00 ₺'
                cell.alignment = styles['number_alignment']
            elif col_idx == 6:  # KDV oranı
                cell.number_format = '0'
                cell.alignment = styles['number_alignment']
        
        total_kdv_haric += kdv_haric
        total_kdv += kdv
        total_indirilen += float(inv.get('toplam_indirilen_kdv', 0) or 0)
    
    # Toplam satırı
    total_row = len(active_invoices) + 5
    ws.cell(row=total_row, column=4, value="TOPLAM").font = Font(bold=True)
    
    total_cells = [
        (5, total_kdv_haric),
        (7, total_kdv),
        (8, total_indirilen)
    ]
    
    for col, value in total_cells:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00 ₺'
        cell.border = styles['thin_border']
    
    # Özet bilgi
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value=f"Toplam Fatura Sayısı: {len(active_invoices)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, color="2E74B5")
    
    # Kaydet
    wb.save(output_stream)
    output_stream.seek(0)
    
    return output_stream


def generate_gib_kalemli_excel(invoices, output_stream=None):
    """
    GİB Kalem Bazlı Excel dosyası oluştur.
    Her kalem = 1 satır.
    
    Args:
        invoices: Fatura listesi (dict listesi)
        output_stream: Opsiyonel BytesIO stream. None ise yeni oluşturulur.
        
    Returns:
        BytesIO: Excel dosyası içeren stream
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli: pip install openpyxl")
    
    if output_stream is None:
        output_stream = io.BytesIO()
    
    # Silinen faturaları filtrele
    active_invoices = [inv for inv in invoices if not inv.get('_deleted', False)]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GİB Kalem Bazlı"
    
    styles = _get_styles()
    
    # Başlık satırı
    ws.merge_cells('A1:K1')
    ws['A1'] = "GİB İNDİRİLECEK KDV LİSTESİ - KALEM BAZLI"
    ws['A1'].font = styles['title_font']
    ws['A1'].alignment = styles['title_alignment']
    
    # Tarih bilgisi
    ws['A2'] = f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=9, color="666666")
    
    # Kolon başlıkları (Satır 4)
    headers = [
        "Fatura No",
        "Fatura Tarihi",
        "Satıcı VKN/TCKN",
        "Mal/Hizmet Kodu",
        "Mal/Hizmet Açıklaması",
        "Miktar",
        "Birim",
        "Birim Fiyat",
        "Satır Tutarı",
        "KDV Oranı (%)",
        "Satır KDV Tutarı"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']
    
    # Kolon genişlikleri
    col_widths = [22, 12, 15, 15, 40, 10, 8, 15, 15, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Veri satırları
    current_row = 5
    total_tutar = 0
    total_kdv = 0
    kalem_sayisi = 0
    
    for inv in active_invoices:
        fatura_no = f"{inv.get('seri', '')}{inv.get('sira_no', '')}"
        fatura_tarihi = inv.get('tarih', '')
        satici_vkn = inv.get('satici_vkn', '')
        
        kalemler = inv.get('kalemler', [])
        
        # Eğer kalem yoksa, fatura özet bilgisiyle tek satır oluştur
        if not kalemler:
            row_data = [
                fatura_no,
                fatura_tarihi,
                satici_vkn,
                "",  # Kod
                inv.get('mal_cinsi', 'MAL/HİZMET'),  # Açıklama
                inv.get('miktar', '1'),  # Miktar
                "AD",  # Birim
                float(inv.get('kdv_haric_tutar', 0) or 0),  # Birim fiyat
                float(inv.get('kdv_haric_tutar', 0) or 0),  # Satır tutarı
                20,  # KDV oranı
                float(inv.get('kdv', 0) or 0)  # KDV tutarı
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = styles['thin_border']
                
                if col_idx in [8, 9, 11]:  # Tutar kolonları
                    cell.number_format = '#,##0.00 ₺'
                    cell.alignment = styles['number_alignment']
                elif col_idx == 10:  # KDV oranı
                    cell.number_format = '0'
                    cell.alignment = styles['number_alignment']
            
            total_tutar += float(inv.get('kdv_haric_tutar', 0) or 0)
            total_kdv += float(inv.get('kdv', 0) or 0)
            kalem_sayisi += 1
            current_row += 1
        else:
            # Her kalem için ayrı satır
            for kalem in kalemler:
                row_data = [
                    fatura_no,
                    fatura_tarihi,
                    satici_vkn,
                    kalem.get('urun_kodu', ''),
                    kalem.get('urun_adi', ''),
                    kalem.get('miktar', 1),
                    kalem.get('birim', 'AD'),
                    float(kalem.get('birim_fiyat', 0) or 0),
                    float(kalem.get('tutar', 0) or 0),
                    kalem.get('kdv_orani', 20),
                    float(kalem.get('kdv_tutari', 0) or 0)
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = styles['thin_border']
                    
                    if col_idx in [8, 9, 11]:  # Tutar kolonları
                        cell.number_format = '#,##0.00 ₺'
                        cell.alignment = styles['number_alignment']
                    elif col_idx == 6:  # Miktar
                        cell.number_format = '#,##0.00'
                        cell.alignment = styles['number_alignment']
                    elif col_idx == 10:  # KDV oranı
                        cell.number_format = '0'
                        cell.alignment = styles['number_alignment']
                
                total_tutar += float(kalem.get('tutar', 0) or 0)
                total_kdv += float(kalem.get('kdv_tutari', 0) or 0)
                kalem_sayisi += 1
                current_row += 1
    
    # Toplam satırı
    total_row = current_row
    ws.cell(row=total_row, column=8, value="TOPLAM").font = Font(bold=True)
    
    total_cells = [
        (9, total_tutar),
        (11, total_kdv)
    ]
    
    for col, value in total_cells:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00 ₺'
        cell.border = styles['thin_border']
    
    # Özet bilgi
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, 
            value=f"Toplam Fatura: {len(active_invoices)} | Toplam Kalem: {kalem_sayisi}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, color="2E74B5")
    
    # Kaydet
    wb.save(output_stream)
    output_stream.seek(0)
    
    return output_stream


# Test
if __name__ == "__main__":
    # Test verisi
    test_invoices = [
        {
            'tarih': '15.01.2025',
            'seri': 'ABC',
            'sira_no': '2025000001',
            'satici_unvan': 'TEST FİRMA A.Ş.',
            'satici_vkn': '1234567890',
            'kdv_haric_tutar': 10000.00,
            'kdv': 2000.00,
            'toplam_indirilen_kdv': 2000.00,
            'kalemler': [
                {
                    'sira': 1,
                    'urun_kodu': 'PRD001',
                    'urun_adi': 'Test Ürün 1',
                    'miktar': 10,
                    'birim': 'AD',
                    'birim_fiyat': 500.00,
                    'tutar': 5000.00,
                    'kdv_orani': 20,
                    'kdv_tutari': 1000.00
                },
                {
                    'sira': 2,
                    'urun_kodu': 'PRD002',
                    'urun_adi': 'Test Ürün 2',
                    'miktar': 5,
                    'birim': 'AD',
                    'birim_fiyat': 1000.00,
                    'tutar': 5000.00,
                    'kdv_orani': 20,
                    'kdv_tutari': 1000.00
                }
            ]
        }
    ]
    
    # Özet test
    ozet_stream = generate_gib_ozet_excel(test_invoices)
    with open('test_gib_ozet.xlsx', 'wb') as f:
        f.write(ozet_stream.read())
    print("✓ Özet Excel oluşturuldu: test_gib_ozet.xlsx")
    
    # Kalemli test
    kalemli_stream = generate_gib_kalemli_excel(test_invoices)
    with open('test_gib_kalemli.xlsx', 'wb') as f:
        f.write(kalemli_stream.read())
    print("✓ Kalemli Excel oluşturuldu: test_gib_kalemli.xlsx")
