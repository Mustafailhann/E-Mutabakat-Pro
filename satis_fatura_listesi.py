"""
Satış Fatura Listesi (Hesaplanan KDV) Oluşturucu
GİB formatında Satış Fatura Listesi Excel dosyası oluşturur.
İndirilecek KDV listesiyle aynı özelliklere sahip.
"""

import zipfile
import xml.etree.ElementTree as ET
import os
import re
from datetime import datetime

# GIB HTML dönüşümü için
try:
    from gib_viewer import transform_invoice_to_html
    GIB_VIEWER_AVAILABLE = True
except ImportError:
    GIB_VIEWER_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# XML Namespaces
NS = {
    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2'
}

# Character limit for text fields (GIB Excel format requires max 72 chars)
CHAR_LIMIT = 72


def parse_invoice_number(inv_no):
    """Fatura numarasından seri ve sıra no ayır."""
    if not inv_no:
        return "", ""
    match = re.match(r'^([A-Za-z]+)(.+)$', inv_no)
    if match:
        return match.group(1).upper(), match.group(2)
    return "", inv_no


def extract_sales_invoice_data(xml_content):
    """
    XML içeriğinden satış faturası verilerini çıkar.
    """
    if isinstance(xml_content, bytes):
        xml_str = xml_content.decode('utf-8', errors='ignore')
    else:
        xml_str = xml_content
    
    # BOM temizle
    if xml_str.startswith('\ufeff'):
        xml_str = xml_str[1:]
    
    root = ET.fromstring(xml_str)
    
    # Invoice number
    inv_no_elem = root.find('.//cbc:ID', NS)
    inv_no = inv_no_elem.text if inv_no_elem is not None else ""
    
    # Invoice date
    inv_date = root.find('.//cbc:IssueDate', NS)
    if inv_date is not None and inv_date.text:
        try:
            dt = datetime.strptime(inv_date.text, '%Y-%m-%d')
            date_formatted = dt.strftime('%d.%m.%Y')
            kdv_period = dt.strftime('%Y/%m')
        except:
            date_formatted = inv_date.text
            kdv_period = ""
    else:
        date_formatted = ""
        kdv_period = ""
    
    # Parse invoice serial and number
    seri, sira_no = parse_invoice_number(inv_no)
    
    # Alıcı (müşteri) bilgileri - Satış faturasında alıcı önemli
    buyer_name = root.find('.//cac:AccountingCustomerParty//cac:PartyName/cbc:Name', NS)
    
    # Alıcı VKN/TCKN
    buyer_vkn = None
    buyer_ids = root.findall('.//cac:AccountingCustomerParty//cac:PartyIdentification/cbc:ID', NS)
    for pid in buyer_ids:
        scheme = pid.get('schemeID', '')
        if scheme in ('VKN', 'TCKN', 'VKN_TCKN'):
            buyer_vkn = pid
            break
        elif not scheme and pid.text:
            if len(pid.text.strip()) in (10, 11) and pid.text.strip().isdigit():
                buyer_vkn = pid
                break
    
    # Satıcı VKN (kendi VKN'miz)
    seller_vkn = None
    seller_ids = root.findall('.//cac:AccountingSupplierParty//cac:PartyIdentification/cbc:ID', NS)
    for pid in seller_ids:
        scheme = pid.get('schemeID', '')
        if scheme in ('VKN', 'TCKN', 'VKN_TCKN'):
            seller_vkn = pid.text if pid.text else None
            break
    
    # Currency and exchange rate
    doc_currency = root.find('.//cbc:DocumentCurrencyCode', NS)
    currency = doc_currency.text if doc_currency is not None else "TRY"
    
    exchange_rate = 1.0
    if currency != "TRY":
        rate_elem = root.find('.//cac:PricingExchangeRate/cbc:CalculationRate', NS)
        if rate_elem is not None:
            try:
                exchange_rate = float(rate_elem.text)
            except:
                pass
    
    # Tax amounts - TL cinsinden
    tax_excl_elem = root.find('.//cac:LegalMonetaryTotal/cbc:TaxExclusiveAmount', NS)
    tax_excl_amount = float(tax_excl_elem.text) if tax_excl_elem is not None and tax_excl_elem.text else 0.0
    
    tax_elem = root.find('.//cac:TaxTotal/cbc:TaxAmount', NS)
    tax_amount = float(tax_elem.text) if tax_elem is not None and tax_elem.text else 0.0
    
    # TL'ye çevir
    if currency != "TRY" and exchange_rate > 0:
        tax_excl_amount = tax_excl_amount * exchange_rate
        tax_amount = tax_amount * exchange_rate
    
    # KDV oranı
    tax_percent = root.find('.//cac:TaxTotal/cac:TaxSubtotal/cac:TaxCategory/cbc:Percent', NS)
    kdv_rate = float(tax_percent.text) if tax_percent is not None and tax_percent.text else 20.0
    
    # Mal/hizmet bilgileri - KALEM BAZLI
    invoice_lines = root.findall('.//cac:InvoiceLine', NS)
    kalemler = []  # YENİ: Her kalem ayrı ayrı saklanacak
    mal_cinsi_parts = []
    miktar_parts = []
    
    for idx, line in enumerate(invoice_lines, 1):
        # Ürün adı
        item_name = line.find('.//cac:Item/cbc:Name', NS)
        urun_adi = item_name.text.strip() if item_name is not None and item_name.text else ""
        
        # Ürün kodu (varsa)
        item_id = line.find('.//cac:Item/cac:SellersItemIdentification/cbc:ID', NS)
        if item_id is None:
            item_id = line.find('.//cac:Item/cac:BuyersItemIdentification/cbc:ID', NS)
        urun_kodu = item_id.text.strip() if item_id is not None and item_id.text else ""
        
        # Miktar ve birim
        qty_elem = line.find('.//cbc:InvoicedQuantity', NS)
        if qty_elem is not None and qty_elem.text:
            try:
                miktar_val = float(qty_elem.text)
            except:
                miktar_val = 1.0
            birim = qty_elem.get('unitCode', 'AD')
        else:
            miktar_val = 1.0
            birim = 'AD'
        
        # Birim fiyat
        price_elem = line.find('.//cac:Price/cbc:PriceAmount', NS)
        if price_elem is not None and price_elem.text:
            try:
                birim_fiyat = float(price_elem.text)
            except:
                birim_fiyat = 0.0
        else:
            birim_fiyat = 0.0
        
        # Satır tutarı (KDV hariç)
        line_ext_elem = line.find('.//cbc:LineExtensionAmount', NS)
        if line_ext_elem is not None and line_ext_elem.text:
            try:
                satir_tutar = float(line_ext_elem.text)
            except:
                satir_tutar = miktar_val * birim_fiyat
        else:
            satir_tutar = miktar_val * birim_fiyat
        
        # Satır KDV oranı ve tutarı
        line_tax_subtotal = line.find('.//cac:TaxTotal/cac:TaxSubtotal', NS)
        if line_tax_subtotal is not None:
            line_tax_percent = line_tax_subtotal.find('.//cac:TaxCategory/cbc:Percent', NS)
            line_tax_amount = line_tax_subtotal.find('.//cbc:TaxAmount', NS)
            try:
                kalem_kdv_orani = float(line_tax_percent.text) if line_tax_percent is not None and line_tax_percent.text else kdv_rate
                kalem_kdv_tutari = float(line_tax_amount.text) if line_tax_amount is not None and line_tax_amount.text else 0.0
            except:
                kalem_kdv_orani = kdv_rate
                kalem_kdv_tutari = satir_tutar * kalem_kdv_orani / 100
        else:
            kalem_kdv_orani = kdv_rate
            kalem_kdv_tutari = satir_tutar * kalem_kdv_orani / 100
        
        # TL'ye çevir
        if currency != "TRY" and exchange_rate > 0:
            satir_tutar = satir_tutar * exchange_rate
            kalem_kdv_tutari = kalem_kdv_tutari * exchange_rate
            birim_fiyat = birim_fiyat * exchange_rate
        
        # Kalem objesi oluştur
        kalem = {
            'sira': idx,
            'urun_kodu': urun_kodu,
            'urun_adi': urun_adi,
            'miktar': miktar_val,
            'birim': birim,
            'birim_fiyat': round(birim_fiyat, 4),
            'tutar': round(satir_tutar, 2),
            'kdv_orani': kalem_kdv_orani,
            'kdv_tutari': round(kalem_kdv_tutari, 2),
            'toplam': round(satir_tutar + kalem_kdv_tutari, 2)
        }
        kalemler.append(kalem)
        
        # Geriye uyumluluk için birleştir
        if urun_adi:
            mal_cinsi_parts.append(urun_adi[:50])
        miktar_parts.append(f"{miktar_val}{birim}")
    
    # 72 karakter limiti ile birleştir (geriye uyumluluk)
    mal_cinsi = "-".join(mal_cinsi_parts)[:CHAR_LIMIT] if mal_cinsi_parts else "MAL/HİZMET"
    miktar = "-".join(miktar_parts)[:CHAR_LIMIT] if miktar_parts else "1AD"
    
    return {
        'tarih': date_formatted,
        'seri': seri,
        'sira_no': sira_no,
        'alici_unvan': buyer_name.text[:CHAR_LIMIT] if buyer_name is not None and buyer_name.text else "",
        'alici_vkn': buyer_vkn.text if buyer_vkn is not None else "",
        'mal_cinsi': mal_cinsi,
        'miktar': miktar,
        'kalemler': kalemler,  # YENİ: Kalem bazlı detay
        'kalem_sayisi': len(kalemler),  # YENİ: Toplam kalem sayısı
        'kdv_haric_tutar': round(tax_excl_amount, 2),
        'kdv_orani': kdv_rate,
        'kdv': round(tax_amount, 2),
        'kdv_donemi': kdv_period,
        'currency': currency,
        'exchange_rate': exchange_rate,
        'seller_vkn': seller_vkn
    }


def load_sales_invoices_from_zip(zip_path, own_vkn=None, period_filter=None):
    """
    ZIP dosyasından satış faturalarını yükle.
    own_vkn: Kendi VKN'miz - bu VKN satıcı olan faturalar satış faturasıdır.
    """
    invoices = []
    
    if not os.path.exists(zip_path):
        print(f"ZIP dosyası bulunamadı: {zip_path}")
        return invoices
    
    # GIB HTML dosyaları için klasör
    gib_html_dir = os.path.join(os.path.dirname(zip_path), "gib_html_satis")
    if GIB_VIEWER_AVAILABLE and not os.path.exists(gib_html_dir):
        os.makedirs(gib_html_dir, exist_ok=True)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            for name in zf.namelist():
                if name.endswith('.xml'):
                    try:
                        xml_data = zf.read(name)
                        inv_data = extract_sales_invoice_data(xml_data)
                        
                        # Sadece kendi satış faturalarımızı al
                        if own_vkn and inv_data.get('seller_vkn') != own_vkn:
                            continue
                        
                        if period_filter and inv_data['kdv_donemi'] != period_filter:
                            continue
                        
                        # GIB HTML oluştur
                        if GIB_VIEWER_AVAILABLE:
                            try:
                                gib_html = transform_invoice_to_html(xml_data.decode('utf-8'))
                                fatura_no = f"{inv_data['seri']}{inv_data['sira_no']}"
                                html_path = os.path.join(gib_html_dir, f"{fatura_no}.html")
                                with open(html_path, 'w', encoding='utf-8') as f:
                                    f.write(gib_html)
                                inv_data['gib_html_path'] = f"file:///{html_path.replace(os.sep, '/')}"
                            except:
                                inv_data['gib_html_path'] = None
                        else:
                            inv_data['gib_html_path'] = None
                        
                        invoices.append(inv_data)
                    except Exception as e:
                        print(f"XML parse hatası ({name}): {e}")
                
                elif name.endswith('.zip'):
                    try:
                        with zf.open(name) as nested:
                            with zipfile.ZipFile(nested) as nzf:
                                for inner in nzf.namelist():
                                    if inner.endswith('.xml'):
                                        try:
                                            xml_data = nzf.read(inner)
                                            inv_data = extract_sales_invoice_data(xml_data)
                                            
                                            if own_vkn and inv_data.get('seller_vkn') != own_vkn:
                                                continue
                                            
                                            if period_filter and inv_data['kdv_donemi'] != period_filter:
                                                continue
                                            
                                            if GIB_VIEWER_AVAILABLE:
                                                try:
                                                    gib_html = transform_invoice_to_html(xml_data.decode('utf-8'))
                                                    fatura_no = f"{inv_data['seri']}{inv_data['sira_no']}"
                                                    html_path = os.path.join(gib_html_dir, f"{fatura_no}.html")
                                                    with open(html_path, 'w', encoding='utf-8') as f:
                                                        f.write(gib_html)
                                                    inv_data['gib_html_path'] = f"file:///{html_path.replace(os.sep, '/')}"
                                                except:
                                                    inv_data['gib_html_path'] = None
                                            else:
                                                inv_data['gib_html_path'] = None
                                            
                                            invoices.append(inv_data)
                                        except Exception as e:
                                            print(f"Fatura parse hatası ({inner}): {e}")
                    except Exception as e:
                        print(f"Nested ZIP hatası ({name}): {e}")
    except zipfile.BadZipFile:
        print(f"Geçersiz ZIP dosyası: {zip_path}")
    
    return invoices


def generate_sales_listesi_excel(invoices, output_path):
    """
    Satış Fatura Listesi Excel dosyası oluştur.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Satış Fatura Listesi"
    
    # Başlıklar
    headers = [
        "Sıra No",
        "Satış Faturasının Tarihi",
        "Satış Faturasının Serisi",
        "Satış Faturasının Sıra No'su",
        "Alıcının Adı-Soyadı / Ünvanı",
        "Alıcının VKN/TCKN",
        "Satılan Mal ve/veya Hizmetin Cinsi",
        "Satılan Mal ve/veya Hizmetin Miktarı",
        "KDV Hariç Tutarı",
        "KDV Oranı %",
        "Hesaplanan KDV",
        "KDV Dönemi"
    ]
    
    # Stil tanımları
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    number_format = '#,##0.00'
    
    # Başlık satırı
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Veri satırları
    for row_idx, inv in enumerate(invoices, 2):
        sira = row_idx - 1
        fatura_no = (inv.get('seri', '') or '') + (inv.get('sira_no', '') or '')
        
        row_data = [
            sira,
            inv.get('tarih', ''),
            '',  # Seri boş
            fatura_no,  # Tam fatura numarası
            inv.get('alici_unvan', '')[:CHAR_LIMIT],
            inv.get('alici_vkn', ''),
            inv.get('mal_cinsi', '')[:CHAR_LIMIT],
            inv.get('miktar', ''),
            inv.get('kdv_haric_tutar', 0),
            inv.get('kdv_orani', 20),
            inv.get('kdv', 0),
            inv.get('kdv_donemi', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in [9, 11]:  # Tutar sütunları
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right")
    
    # Toplam satırı
    total_row = len(invoices) + 2
    ws.cell(row=total_row, column=7, value="TOPLAM").font = Font(bold=True)
    
    total_kdv_haric = sum(inv.get('kdv_haric_tutar', 0) for inv in invoices)
    total_kdv = sum(inv.get('kdv', 0) for inv in invoices)
    
    ws.cell(row=total_row, column=9, value=total_kdv_haric).number_format = number_format
    ws.cell(row=total_row, column=11, value=total_kdv).number_format = number_format
    
    for col in range(1, 13):
        ws.cell(row=total_row, column=col).border = thin_border
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    
    # Sütun genişlikleri
    col_widths = [8, 15, 10, 22, 40, 15, 50, 20, 18, 10, 18, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    print("Satış Fatura Listesi modülü yüklendi.")
