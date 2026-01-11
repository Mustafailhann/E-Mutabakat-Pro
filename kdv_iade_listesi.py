"""
KDV İade Listesi Oluşturucu
GİB'in resmi İndirilecek KDV Listesi formatında Excel dosyası oluşturur.
"""
import zipfile
import xml.etree.ElementTree as ET
import os
import re
from datetime import datetime

# GIB HTML dönüşümü için
try:
    from gib_viewer import transform_invoice_to_html
    GIB_VIEWER_AVAILABLE = True
except ImportError:
    GIB_VIEWER_AVAILABLE = False

try:
    import xlwt
    XLWT_AVAILABLE = True
except ImportError:
    XLWT_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# XML Namespaces
NS = {
    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2'
}

# Character limit for text fields (GIB Excel format requires max 72 chars)
CHAR_LIMIT = 72

# Unit code mappings for Turkish
UNIT_MAP = {
    'C62': 'AD',  # Adet
    'KGM': 'KG',  # Kilogram
    'LTR': 'LT',  # Litre
    'MTR': 'MT',  # Metre
    'MTK': 'M2',  # Metrekare
    'MTQ': 'M3',  # Metreküp
    'KWH': 'KWH', # Kilowatt-saat
    'TNE': 'TON', # Ton
    'NIU': 'AD',  # Number of units
    'PR': 'ÇFT',  # Pair (çift)
    'SET': 'SET', # Set
    'BX': 'KT',   # Box (kutu)
    'PK': 'PK',   # Package (paket)
}


def parse_invoice_number(inv_no):
    """
    Fatura numarasından seri ve sıra numarasını ayır.
    Örnek: ZGM2025000001473 -> ('ZGM', '2025000001473')
    """
    if not inv_no:
        return '', ''
    
    # Find where letters end and numbers begin
    match = re.match(r'^([A-Za-z]+)(.+)$', inv_no)
    if match:
        return match.group(1), match.group(2)
    return '', inv_no


def format_quantity(qty, unit_code):
    """
    Miktar ve birimi formatla.
    Örnek: (100, 'KGM') -> '100KG'
    """
    if not qty:
        return ''
    
    # Clean quantity
    try:
        qty_val = float(qty)
        if qty_val == int(qty_val):
            qty_str = str(int(qty_val))
        else:
            qty_str = f"{qty_val:.2f}".rstrip('0').rstrip('.')
    except:
        qty_str = str(qty)
    
    # Map unit
    unit = UNIT_MAP.get(unit_code, unit_code if unit_code else 'AD')
    
    return f"{qty_str}{unit}"


def truncate_with_items(items_list, max_length=CHAR_LIMIT, separator='-'):
    """
    Birden fazla kalemi tire ile birleştir, 255 karakteri aşarsa kısalt.
    """
    if not items_list:
        return ''
    
    # Clean and filter empty items
    items_list = [str(item).strip() for item in items_list if item and str(item).strip()]
    
    if not items_list:
        return ''
    
    # Try to fit all items
    result = separator.join(items_list)
    
    if len(result) <= max_length:
        return result
    
    # Need to truncate - try to include as many items as possible
    result_items = []
    current_length = 0
    
    for item in items_list:
        # Calculate length if we add this item
        if result_items:
            needed = len(separator) + len(item)
        else:
            needed = len(item)
        
        if current_length + needed <= max_length - 3:  # Leave room for "..."
            result_items.append(item)
            current_length += needed
        else:
            # Can't fit more - truncate last item if needed
            break
    
    if result_items:
        result = separator.join(result_items)
        if len(result) > max_length:
            result = result[:max_length-3] + '...'
    else:
        # First item alone is too long
        result = items_list[0][:max_length-3] + '...'
    
    return result


def extract_invoice_data(xml_content):
    """
    E-fatura XML'inden KDV listesi için gerekli verileri çıkar.
    Dövizli faturalar TL'ye çevrilir.
    """
    root = ET.fromstring(xml_content)
    
    # Basic invoice info
    inv_id = root.find('.//cbc:ID', NS)
    inv_no = inv_id.text if inv_id is not None else ""
    
    issue_date = root.find('.//cbc:IssueDate', NS)
    date_str = issue_date.text if issue_date is not None else ""
    
    # Format date as DD.MM.YYYY
    if date_str:
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            date_formatted = dt.strftime('%d.%m.%Y')
            kdv_period = dt.strftime('%Y/%m')
        except:
            date_formatted = date_str
            kdv_period = ""
    else:
        date_formatted = ""
        kdv_period = ""
    
    # Parse invoice serial and number
    seri, sira_no = parse_invoice_number(inv_no)
    
    # Supplier info - önce şirket adını dene, yoksa şahıs adını kontrol et
    supplier_name = root.find('.//cac:AccountingSupplierParty//cac:PartyName/cbc:Name', NS)
    
    # Şahıs faturası: PartyName boşsa Person/FirstName + FamilyName kullan
    supplier_name_text = ""
    if supplier_name is not None and supplier_name.text:
        supplier_name_text = supplier_name.text.strip()
    
    if not supplier_name_text:
        # Şahıs adını kontrol et
        supplier_party = root.find('.//cac:AccountingSupplierParty/cac:Party', NS)
        if supplier_party is not None:
            first_name = supplier_party.find('cac:Person/cbc:FirstName', NS)
            family_name = supplier_party.find('cac:Person/cbc:FamilyName', NS)
            
            name_parts = []
            if first_name is not None and first_name.text:
                name_parts.append(first_name.text.strip())
            if family_name is not None and family_name.text:
                name_parts.append(family_name.text.strip())
            
            if name_parts:
                supplier_name_text = ' '.join(name_parts)
    
    # VKN/TCKN - filter by schemeID (not phone numbers)
    supplier_vkn = None
    party_ids = root.findall('.//cac:AccountingSupplierParty//cac:PartyIdentification/cbc:ID', NS)
    for pid in party_ids:
        scheme = pid.get('schemeID', '')
        # VKN (10 haneli) veya TCKN (11 haneli) olmalı
        if scheme in ('VKN', 'TCKN', 'VKN_TCKN'):
            supplier_vkn = pid
            break
        # schemeID yoksa ID uzunluğuna bak (VKN=10, TCKN=11)
        elif not scheme and pid.text:
            if len(pid.text.strip()) in (10, 11) and pid.text.strip().isdigit():
                supplier_vkn = pid
                break
    
    # Alıcı (müşteri) VKN - satış faturası kontrolü için
    buyer_vkn = None
    buyer_ids = root.findall('.//cac:AccountingCustomerParty//cac:PartyIdentification/cbc:ID', NS)
    for pid in buyer_ids:
        scheme = pid.get('schemeID', '')
        if scheme in ('VKN', 'TCKN', 'VKN_TCKN'):
            buyer_vkn = pid.text if pid.text else None
            break
        elif not scheme and pid.text:
            if len(pid.text.strip()) in (10, 11) and pid.text.strip().isdigit():
                buyer_vkn = pid.text.strip()
                break
    
    # Currency and exchange rate
    doc_currency = root.find('.//cbc:DocumentCurrencyCode', NS)
    currency = doc_currency.text if doc_currency is not None else "TRY"
    
    # Get exchange rate
    exchange_rate = 1.0
    if currency != "TRY":
        # Try to find exchange rate in different locations
        rate_elem = root.find('.//cac:PricingExchangeRate/cbc:CalculationRate', NS)
        if rate_elem is not None:
            try:
                exchange_rate = float(rate_elem.text)
            except:
                pass
        
        # Alternative: look in PayableAmount with TRY
        if exchange_rate == 1.0:
            payable_try = root.find('.//cac:LegalMonetaryTotal/cbc:PayableAmount[@currencyID="TRY"]', NS)
            payable_orig = root.find('.//cac:LegalMonetaryTotal/cbc:PayableAmount', NS)
            if payable_try is not None and payable_orig is not None:
                try:
                    orig_amount = float(payable_orig.text)
                    if orig_amount > 0:
                        # Exchange rate might be in a note
                        pass
                except:
                    pass
    
    # Tax totals
    tax_total = root.find('.//cac:TaxTotal/cbc:TaxAmount', NS)
    tax_amount = float(tax_total.text) if tax_total is not None else 0.0
    
    # Tax exclusive amount
    tax_excl = root.find('.//cbc:TaxExclusiveAmount', NS)
    if tax_excl is None:
        tax_excl = root.find('.//cac:LegalMonetaryTotal/cbc:TaxExclusiveAmount', NS)
    tax_excl_amount = float(tax_excl.text) if tax_excl is not None else 0.0
    
    # Convert to TL if foreign currency
    if currency != "TRY" and exchange_rate > 1.0:
        tax_amount = tax_amount * exchange_rate
        tax_excl_amount = tax_excl_amount * exchange_rate
    
    # Withholding tax (tevkifat)
    withholding = root.find('.//cac:WithholdingTaxTotal/cbc:TaxAmount', NS)
    withholding_amount = float(withholding.text) if withholding is not None else 0.0
    if currency != "TRY" and exchange_rate > 1.0:
        withholding_amount = withholding_amount * exchange_rate
    
    # Collect item names and quantities - KALEM BAZLI
    kalemler = []  # YENİ: Her kalem ayrı ayrı saklanacak
    item_names = []
    item_quantities = []
    
    lines = root.findall('.//cac:InvoiceLine', NS)
    for idx, line in enumerate(lines, 1):
        # Ürün adı
        item_name = line.find('.//cac:Item/cbc:Name', NS)
        if item_name is not None and item_name.text:
            urun_adi = item_name.text.strip()
        else:
            # Fallback to description
            item_desc = line.find('.//cac:Item/cbc:Description', NS)
            if item_desc is not None and item_desc.text:
                urun_adi = item_desc.text.strip()
            else:
                urun_adi = "MAL/HİZMET"
        
        # Clean up name - remove extra whitespace, newlines
        urun_adi = ' '.join(urun_adi.split())
        
        # Ürün kodu (varsa)
        item_id = line.find('.//cac:Item/cac:SellersItemIdentification/cbc:ID', NS)
        if item_id is None:
            item_id = line.find('.//cac:Item/cac:BuyersItemIdentification/cbc:ID', NS)
        urun_kodu = item_id.text.strip() if item_id is not None and item_id.text else ""
        
        # Miktar ve birim
        qty_elem = line.find('cbc:InvoicedQuantity', NS)
        if qty_elem is not None and qty_elem.text:
            try:
                miktar_val = float(qty_elem.text)
            except:
                miktar_val = 1.0
            birim_code = qty_elem.get('unitCode', 'C62')
            birim = UNIT_MAP.get(birim_code, birim_code if birim_code else 'AD')
        else:
            miktar_val = 1.0
            birim = 'AD'
        
        # Birim fiyat
        price_elem = line.find('.//cac:Price/cbc:PriceAmount', NS)
        if price_elem is not None and price_elem.text:
            try:
                birim_fiyat = float(price_elem.text)
            except:
                birim_fiyat = 0.0
        else:
            birim_fiyat = 0.0
        
        # Satır tutarı (KDV hariç)
        line_ext_elem = line.find('.//cbc:LineExtensionAmount', NS)
        if line_ext_elem is not None and line_ext_elem.text:
            try:
                satir_tutar = float(line_ext_elem.text)
            except:
                satir_tutar = miktar_val * birim_fiyat
        else:
            satir_tutar = miktar_val * birim_fiyat
        
        # Satır KDV oranı ve tutarı
        line_tax_subtotal = line.find('.//cac:TaxTotal/cac:TaxSubtotal', NS)
        if line_tax_subtotal is not None:
            line_tax_percent = line_tax_subtotal.find('.//cac:TaxCategory/cbc:Percent', NS)
            line_tax_amount_elem = line_tax_subtotal.find('.//cbc:TaxAmount', NS)
            try:
                kalem_kdv_orani = float(line_tax_percent.text) if line_tax_percent is not None and line_tax_percent.text else 20.0
                kalem_kdv_tutari = float(line_tax_amount_elem.text) if line_tax_amount_elem is not None and line_tax_amount_elem.text else 0.0
            except:
                kalem_kdv_orani = 20.0
                kalem_kdv_tutari = satir_tutar * kalem_kdv_orani / 100
        else:
            kalem_kdv_orani = 20.0
            kalem_kdv_tutari = satir_tutar * kalem_kdv_orani / 100
        
        # TL'ye çevir
        if currency != "TRY" and exchange_rate > 1.0:
            satir_tutar = satir_tutar * exchange_rate
            kalem_kdv_tutari = kalem_kdv_tutari * exchange_rate
            birim_fiyat = birim_fiyat * exchange_rate
        
        # Kalem objesi oluştur
        kalem = {
            'sira': idx,
            'urun_kodu': urun_kodu,
            'urun_adi': urun_adi,
            'miktar': miktar_val,
            'birim': birim,
            'birim_fiyat': round(birim_fiyat, 4),
            'tutar': round(satir_tutar, 2),
            'kdv_orani': kalem_kdv_orani,
            'kdv_tutari': round(kalem_kdv_tutari, 2),
            'toplam': round(satir_tutar + kalem_kdv_tutari, 2)
        }
        kalemler.append(kalem)
        
        # Geriye uyumluluk için birleştir
        item_names.append(urun_adi)
        qty_formatted = format_quantity(miktar_val, birim)
        item_quantities.append(qty_formatted)
    
    # Combine items with 72 char limit (geriye uyumluluk)
    mal_cinsi = truncate_with_items(item_names, CHAR_LIMIT,  '-')
    miktar = truncate_with_items(item_quantities, CHAR_LIMIT, '-')
    
    # If no items found, use generic
    if not mal_cinsi:
        mal_cinsi = "MAL/HİZMET"
    if not miktar:
        miktar = "1AD"
    
    # Calculate tevkifat-related KDV
    # Tevkifatsız KDV = KDV - Tevkifat (Tevkifata Tabi Olmayan KDV)
    tevkifat_tabi_olmayan_kdv = 0.0
    iki_nolu_kdv = 0.0
    
    if withholding_amount > 0:
        # Tevkifatlı fatura
        # Tevkifatsız KDV = KDV - Tevkifat
        tevkifat_tabi_olmayan_kdv = round(tax_amount - withholding_amount, 2)
        # 2 No KDV = KDV - Tevkifatsız KDV = Tevkifat tutarı
        iki_nolu_kdv = round(tax_amount - tevkifat_tabi_olmayan_kdv, 2)
    else:
        # Tevkifatsız fatura - tüm KDV doğrudan indirilebilir
        tevkifat_tabi_olmayan_kdv = round(tax_amount, 2)
        iki_nolu_kdv = 0.0
    
    # Toplam indirilen KDV = Tevkifatsız KDV + 2 No KDV
    toplam_indirilen = round(tevkifat_tabi_olmayan_kdv + iki_nolu_kdv, 2)
    
    return {
        'tarih': date_formatted,
        'seri': seri,
        'sira_no': sira_no,
        'satici_unvan': supplier_name_text,
        'satici_vkn': supplier_vkn.text if supplier_vkn is not None else "",
        'mal_cinsi': mal_cinsi,
        'miktar': miktar,
        'kalemler': kalemler,  # YENİ: Kalem bazlı detay
        'kalem_sayisi': len(kalemler),  # YENİ: Toplam kalem sayısı
        'kdv_haric_tutar': round(tax_excl_amount, 2),  # TL cinsinden
        'kdv': round(tax_amount, 2),  # TL cinsinden - Toplam KDV
        'tevkifat_kdv': tevkifat_tabi_olmayan_kdv,  # Tevkifatsız KDV = KDV - Tevkifat
        'iki_nolu_kdv': iki_nolu_kdv,  # 2 No KDV = KDV - Tevkifatsız KDV
        'toplam_indirilen_kdv': toplam_indirilen,  # İndirilen KDV = Tevkifatsız KDV + 2 No KDV
        'ggb_tescil_no': "",  # İthalat değilse boş
        'kdv_donemi': kdv_period,
        'currency': currency,
        'exchange_rate': exchange_rate,
        'withholding_amount': round(withholding_amount, 2),  # Tevkifat tutarı (debug için)
        'buyer_vkn': buyer_vkn  # Alıcı VKN - satış faturası kontrolü için
    }


def load_invoices_from_zip(zip_path, period_filter=None):
    """
    ZIP dosyasından faturaları yükle.
    Hem nested ZIP yapısını hem de doğrudan XML içeren ZIP'leri destekler.
    period_filter: 'YYYY/MM' formatında dönem filtresi (opsiyonel)
    """
    invoices = []
    
    if not os.path.exists(zip_path):
        print(f"ZIP dosyası bulunamadı: {zip_path}")
        return invoices
    
    # GIB HTML dosyaları için klasör oluştur
    gib_html_dir = os.path.join(os.path.dirname(zip_path), "gib_html")
    if GIB_VIEWER_AVAILABLE and not os.path.exists(gib_html_dir):
        os.makedirs(gib_html_dir, exist_ok=True)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            for name in zf.namelist():
                # Doğrudan XML dosyası (tekil fatura ZIP'i)
                if name.endswith('.xml'):
                    try:
                        xml_data = zf.read(name)
                        inv_data = extract_invoice_data(xml_data)
                        
                        if period_filter and inv_data['kdv_donemi'] != period_filter:
                            continue
                        
                        # Save original XML file to disk
                        fatura_no = f"{inv_data['seri']}{inv_data['sira_no']}"
                        xml_dir = os.path.join(os.path.dirname(zip_path), "xml_files")
                        if not os.path.exists(xml_dir):
                            os.makedirs(xml_dir, exist_ok=True)
                        xml_filename = f"{fatura_no}.xml"
                        xml_path = os.path.join(xml_dir, xml_filename)
                        with open(xml_path, 'wb') as f:
                            f.write(xml_data)
                        inv_data['source_path'] = xml_path.replace(os.sep, '/')
                        
                        # GIB HTML oluştur ve kaydet
                        if GIB_VIEWER_AVAILABLE:
                            try:
                                gib_html, _ = transform_invoice_to_html(xml_data.decode('utf-8'))
                                fatura_no = f"{inv_data['seri']}{inv_data['sira_no']}"
                                html_filename = f"{fatura_no}.html"
                                html_path = os.path.join(gib_html_dir, html_filename)
                                with open(html_path, 'w', encoding='utf-8') as f:
                                    f.write(gib_html)
                                inv_data['gib_html_path'] = f"file:///{html_path.replace(os.sep, '/')}"
                            except Exception as e:
                                print(f"GIB HTML hatası ({name}): {e}")
                                inv_data['gib_html_path'] = None
                        else:
                            inv_data['gib_html_path'] = None
                        
                        invoices.append(inv_data)
                    except Exception as e:
                        print(f"XML parse hatası ({name}): {e}")
                
                # Nested ZIP (ana arşiv yapısı)
                elif name.endswith('.zip'):
                    try:
                        with zf.open(name) as nested:
                            with zipfile.ZipFile(nested) as nzf:
                                for inner in nzf.namelist():
                                    if inner.endswith('.xml'):
                                        try:
                                            xml_data = nzf.read(inner)
                                            inv_data = extract_invoice_data(xml_data)
                                            
                                            if period_filter and inv_data['kdv_donemi'] != period_filter:
                                                continue
                                            
                                            # Save original XML file to disk
                                            fatura_no = f"{inv_data['seri']}{inv_data['sira_no']}"
                                            xml_dir = os.path.join(os.path.dirname(zip_path), "xml_files")
                                            if not os.path.exists(xml_dir):
                                                os.makedirs(xml_dir, exist_ok=True)
                                            xml_filename = f"{fatura_no}.xml"
                                            xml_path = os.path.join(xml_dir, xml_filename)
                                            with open(xml_path, 'wb') as f:
                                                f.write(xml_data)
                                            inv_data['source_path'] = xml_path.replace(os.sep, '/')
                                            
                                            # GIB HTML oluştur ve kaydet
                                            if GIB_VIEWER_AVAILABLE:
                                                try:
                                                    gib_html, _ = transform_invoice_to_html(xml_data.decode('utf-8'))
                                                    fatura_no = f"{inv_data['seri']}{inv_data['sira_no']}"
                                                    html_filename = f"{fatura_no}.html"
                                                    html_path = os.path.join(gib_html_dir, html_filename)
                                                    with open(html_path, 'w', encoding='utf-8') as f:
                                                        f.write(gib_html)
                                                    inv_data['gib_html_path'] = f"file:///{html_path.replace(os.sep, '/')}"
                                                except Exception as e:
                                                    print(f"GIB HTML hatası ({inner}): {e}")
                                                    inv_data['gib_html_path'] = None
                                            else:
                                                inv_data['gib_html_path'] = None
                                            
                                            invoices.append(inv_data)
                                        except Exception as e:
                                            print(f"Fatura parse hatası ({inner}): {e}")
                    except Exception as e:
                        print(f"Nested ZIP hatası ({name}): {e}")
    except zipfile.BadZipFile:
        print(f"Geçersiz ZIP dosyası: {zip_path}")
    
    return invoices


def load_invoice_from_xml(xml_path, period_filter=None):
    """
    Tekil XML dosyasından fatura yükle.
    """
    if not os.path.exists(xml_path):
        print(f"XML dosyası bulunamadı: {xml_path}")
        return None
    
    try:
        with open(xml_path, 'rb') as f:
            xml_data = f.read()
        
        inv_data = extract_invoice_data(xml_data)
        
        # Apply period filter if specified
        if period_filter:
            if inv_data['kdv_donemi'] != period_filter:
                return None
        
        return inv_data
    except Exception as e:
        print(f"XML parse hatası ({xml_path}): {e}")
        return None


def generate_kdv_listesi_excel(invoices, output_path):
    """
    İndirilecek KDV Listesi Excel dosyası oluştur.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İndirilecek KDV Listesi"
    
    # Header row styling
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:O1')
    ws['A1'] = "İNDİRİLECEK KDV LİSTESİ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Empty row
    ws.append([])
    
    # Headers (Row 3)
    headers = [
        "Sıra No",
        "Alış Faturasının Tarihi",
        "Alış Faturasının Serisi",
        "Alış Faturasının Sıra No'su",
        "Satıcının Adı-Soyadı / Ünvanı",
        "Satıcının Vergi Kimlik Numarası / TC Kimlik Numarası",
        "Alınan Mal ve/veya Hizmetin Cinsi",
        "Alınan Mal ve/veya Hizmetin Miktarı",
        "Alınan Mal ve/veya Hizmetin KDV Hariç Tutarı",
        "KDV'si",
        "Tevkifatlı Faturanın Tevkifata Tabi Olmayan Ve Bu Dönemde İndirilen KDV Tutarı",
        "2 Nolu Beyannamede Ödenen KDV Tutarı",
        "Toplam İndirilen KDV Tutarı",
        "GGB Tescil No'su (Alış İthalat İse)",
        "Belgenin İndirim Hakkının Kullanıldığı KDV Dönemi"
    ]
    
    ws.append(headers)
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Column widths
    col_widths = [8, 15, 10, 20, 40, 15, 50, 25, 20, 15, 20, 15, 15, 20, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Data rows
    total_kdv_haric = 0
    total_kdv = 0
    total_indirilen = 0
    
    for idx, inv in enumerate(invoices, 1):
        row = [
            idx,
            inv['tarih'],
            inv['seri'],
            inv['sira_no'],
            inv['satici_unvan'],
            inv['satici_vkn'],
            inv['mal_cinsi'],
            inv['miktar'],
            inv['kdv_haric_tutar'],
            inv['kdv'],
            inv['tevkifat_kdv'],
            inv['iki_nolu_kdv'],
            inv['toplam_indirilen_kdv'],
            inv['ggb_tescil_no'],
            inv['kdv_donemi']
        ]
        ws.append(row)
        
        # Style data cells
        for col in range(1, 16):
            cell = ws.cell(row=idx+3, column=col)
            cell.border = thin_border
            if col in [9, 10, 11, 12, 13]:  # Numeric columns
                cell.number_format = '#,##0.00'
        
        total_kdv_haric += inv['kdv_haric_tutar']
        total_kdv += inv['kdv']
        total_indirilen += inv['toplam_indirilen_kdv']
    
    # Totals row
    total_row = len(invoices) + 4
    ws.cell(row=total_row, column=8, value="TOPLAM")
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=total_kdv_haric)
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    ws.cell(row=total_row, column=10, value=total_kdv)
    ws.cell(row=total_row, column=10).number_format = '#,##0.00'
    ws.cell(row=total_row, column=13, value=total_indirilen)
    ws.cell(row=total_row, column=13).number_format = '#,##0.00'
    
    # Save
    wb.save(output_path)
    print(f"KDV listesi oluşturuldu: {output_path}")
    print(f"Toplam fatura: {len(invoices)}")
    print(f"Toplam KDV Hariç: {total_kdv_haric:,.2f} TL")
    print(f"Toplam KDV: {total_kdv:,.2f} TL")
    
    return output_path


# Test
if __name__ == "__main__":
    work_dir = r"c:\Users\Asus\Desktop\agent ff"
    zip_path = os.path.join(work_dir, "Gelen e-Fatura.zip")
    
    print("Faturalar yükleniyor...")
    invoices = load_invoices_from_zip(zip_path)
    
    print(f"\nToplam {len(invoices)} fatura bulundu.")
    
    if invoices:
        # Show first 3 invoices
        print("\nİlk 3 fatura örneği:")
        for inv in invoices[:3]:
            print(f"  {inv['tarih']} - {inv['seri']}{inv['sira_no']}")
            print(f"    Satıcı: {inv['satici_unvan'][:50]}...")
            print(f"    Mal/Hizmet: {inv['mal_cinsi'][:60]}...")
            print(f"    Miktar: {inv['miktar'][:40]}")
            print(f"    KDV: {inv['kdv']:,.2f}")
        
        # Generate Excel
        output_file = os.path.join(work_dir, "Indirilecek_KDV_Listesi.xlsx")
        generate_kdv_listesi_excel(invoices, output_file)
